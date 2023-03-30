#!/usr/bin/env python3

# Mimic the `gspread` API for a local XLSX file.  This is more convenient for 
# development, because it's much faster and doesn't run into issues with 
# getting throttled by Google.

from openpyxl import load_workbook
from plum import dispatch
from more_itertools import always_iterable
import re

class Document:

    def __init__(self, xlsx_in, xlsx_out=None):
        self.doc = load_workbook(xlsx_in)
        self.output_path = xlsx_out or xlsx_in

    def worksheet(self, name):
        return Worksheet(self, self.doc[name])

    def commit(self):
        self.doc.save(self.output_path)


class Worksheet:

    def __init__(self, doc, sheet):
        self.doc = doc
        self.sheet = sheet

    def get_range(self, range, include_empty=False):
        # The only way to iterate through a range of cells is `sheet[range]`.  
        # But this doesn't return consistent dimensions: If a column range is 
        # given, you get lists of columns.  If a row range is given, you get 
        # lists of rows.  If the range has only one row or one column, you get 
        # only a single list instead of a list-of-lists.  A lot of empty cells 
        # are included as well.  Here, I try to account for all of this and 
        # return a row-wise list-of-lists.
        table = {}
        range, subset = parse_range(range)

        for cells in always_iterable(self.sheet[range]):
            for cell in always_iterable(cells):
                i, j = cell.row - 1, cell.column - 1
                table.setdefault(i, {})[j] = cell

        rows = [
                [cell for (_, cell) in sorted(row.items())]
                for (_, row) in sorted(table.items())
                if not all(cell.value is None for cell in row.values()) \
                        or include_empty
        ]
        return rows[subset]

    def get_values(self, range=None, value_render_option=None):
        if range is None:
            return [
                    row
                    for row in self.sheet.values
                    if not all(x is None for x in row)

            ]

        else:
            return [
                    [cell.value for cell in row]
                    for row in self.get_range(range)
            ]

    def acell(self, cell, value_render_option=None):
        # Return value needs to have a `value` attribute, which `sheet[cell]`
        # coincidentally does.
        return self.sheet[cell]

    @dispatch
    def update(self, range, values):
        if ':' not in range:
            self.sheet[range].value = values
        else:
            cells = self.get_range(range, include_empty=True)
            for i, row in enumerate(cells):
                if i < len(values):
                    for j, cell in enumerate(row):
                        cell.value = values[i][j]
        self.doc.commit()

    @dispatch
    def update(self, values):
        for i, row in enumerate(values, 1):
            for j, value in enumerate(row, 1):
                self.sheet.cell(i, j).value = value
        self.doc.commit()

    def batch_clear(self, range):
        for row in self.get_range(range):
            for cell in row:
                cell.value = None

    def clear(self):
        for col in self.sheet.iter_cols():
            for cell in col:
                cell.value = None

def load_doc(xlsx_in, xlsx_out=None):
    return Document(xlsx_in, xlsx_out)

def parse_range(range):
    # Openpyxl doesn't understand the 'D2:D' syntax, so this function looks for 
    # this and returns (i) a range string that openpyxl will understand and 
    # (ii) a slice that can be applied to the aforementioned range string to 
    # get only the desired rows.
    #
    # There are probably other incompatible range patterns, but I think this is 
    # the only one we actually use.

    if isinstance(range, list):
        assert len(range) == 1
        range = range[0]

    if m := re.match(r'([A-Z])(\d+):([A-Z])', range):
        range = f'{m.group(1)}:{m.group(3)}'
        head = int(m.group(2))
        subset = slice(head - 1, None)

    else:
        subset = slice(None, None)

    return range, subset


